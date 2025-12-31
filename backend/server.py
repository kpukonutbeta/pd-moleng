from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from jose import JWTError, jwt
from passlib.context import CryptContext
import io

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
SECRET_KEY = os.environ.get('JWT_SECRET', 'laporan-perjalanan-dinas-secret-key-2024')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ============ MODELS ============

class UserRegister(BaseModel):
    email: EmailStr
    password: str
    full_name: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserProfile(BaseModel):
    full_name: str
    nip: Optional[str] = ""
    jabatan: Optional[str] = ""
    unit: Optional[str] = ""

class UserResponse(BaseModel):
    id: str
    email: str
    full_name: str
    nip: str
    jabatan: str
    unit: str
    profile_completed: bool

class TripCreate(BaseModel):
    judul: str
    tujuan: str
    tanggal_mulai: str
    tanggal_selesai: str
    dasar_perjalanan: str
    maksud_tujuan: str

class TripUpdate(BaseModel):
    judul: Optional[str] = None
    tujuan: Optional[str] = None
    tanggal_mulai: Optional[str] = None
    tanggal_selesai: Optional[str] = None
    dasar_perjalanan: Optional[str] = None
    maksud_tujuan: Optional[str] = None
    status: Optional[str] = None

class TripResponse(BaseModel):
    id: str
    user_id: str
    judul: str
    tujuan: str
    tanggal_mulai: str
    tanggal_selesai: str
    dasar_perjalanan: str
    maksud_tujuan: str
    status: str
    created_at: str

class ItineraryCreate(BaseModel):
    tanggal: str
    waktu: str
    kegiatan: str
    lokasi: str
    catatan: Optional[str] = ""

class ItineraryUpdate(BaseModel):
    tanggal: Optional[str] = None
    waktu: Optional[str] = None
    kegiatan: Optional[str] = None
    lokasi: Optional[str] = None
    catatan: Optional[str] = None

class ItineraryResponse(BaseModel):
    id: str
    trip_id: str
    tanggal: str
    waktu: str
    kegiatan: str
    lokasi: str
    catatan: str

class ExpenseCreate(BaseModel):
    tanggal: str
    uraian: str
    jumlah: float
    catatan: Optional[str] = ""

class ExpenseUpdate(BaseModel):
    tanggal: Optional[str] = None
    uraian: Optional[str] = None
    jumlah: Optional[float] = None
    catatan: Optional[str] = None

class ExpenseResponse(BaseModel):
    id: str
    trip_id: str
    nomor: int
    tanggal: str
    uraian: str
    jumlah: float
    catatan: str

# ============ AUTH HELPERS ============

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Token tidak valid",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user is None:
        raise credentials_exception
    return user

# ============ AUTH ROUTES ============

@api_router.post("/auth/register")
async def register(user: UserRegister):
    existing = await db.users.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "email": user.email,
        "password": get_password_hash(user.password),
        "full_name": user.full_name,
        "nip": "",
        "jabatan": "",
        "unit": "",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    token = create_access_token({"sub": user_id})
    return {"token": token, "user": {
        "id": user_id,
        "email": user.email,
        "full_name": user.full_name,
        "nip": "",
        "jabatan": "",
        "unit": "",
        "profile_completed": False
    }}

@api_router.post("/auth/login")
async def login(user: UserLogin):
    db_user = await db.users.find_one({"email": user.email}, {"_id": 0})
    if not db_user or not verify_password(user.password, db_user["password"]):
        raise HTTPException(status_code=401, detail="Email atau password salah")
    
    token = create_access_token({"sub": db_user["id"]})
    profile_completed = bool(db_user.get("nip") and db_user.get("jabatan") and db_user.get("unit"))
    return {"token": token, "user": {
        "id": db_user["id"],
        "email": db_user["email"],
        "full_name": db_user["full_name"],
        "nip": db_user.get("nip", ""),
        "jabatan": db_user.get("jabatan", ""),
        "unit": db_user.get("unit", ""),
        "profile_completed": profile_completed
    }}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    profile_completed = bool(current_user.get("nip") and current_user.get("jabatan") and current_user.get("unit"))
    return {
        "id": current_user["id"],
        "email": current_user["email"],
        "full_name": current_user["full_name"],
        "nip": current_user.get("nip", ""),
        "jabatan": current_user.get("jabatan", ""),
        "unit": current_user.get("unit", ""),
        "profile_completed": profile_completed
    }

@api_router.put("/auth/profile")
async def update_profile(profile: UserProfile, current_user: dict = Depends(get_current_user)):
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {
            "full_name": profile.full_name,
            "nip": profile.nip,
            "jabatan": profile.jabatan,
            "unit": profile.unit
        }}
    )
    profile_completed = bool(profile.nip and profile.jabatan and profile.unit)
    return {
        "id": current_user["id"],
        "email": current_user["email"],
        "full_name": profile.full_name,
        "nip": profile.nip,
        "jabatan": profile.jabatan,
        "unit": profile.unit,
        "profile_completed": profile_completed
    }

# ============ TRIP ROUTES ============

@api_router.post("/trips", response_model=TripResponse)
async def create_trip(trip: TripCreate, current_user: dict = Depends(get_current_user)):
    trip_id = str(uuid.uuid4())
    trip_doc = {
        "id": trip_id,
        "user_id": current_user["id"],
        "judul": trip.judul,
        "tujuan": trip.tujuan,
        "tanggal_mulai": trip.tanggal_mulai,
        "tanggal_selesai": trip.tanggal_selesai,
        "dasar_perjalanan": trip.dasar_perjalanan,
        "maksud_tujuan": trip.maksud_tujuan,
        "status": "draft",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.trips.insert_one(trip_doc)
    return TripResponse(**trip_doc)

@api_router.get("/trips", response_model=List[TripResponse])
async def get_trips(current_user: dict = Depends(get_current_user)):
    trips = await db.trips.find({"user_id": current_user["id"]}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [TripResponse(**t) for t in trips]

@api_router.get("/trips/{trip_id}", response_model=TripResponse)
async def get_trip(trip_id: str, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id, "user_id": current_user["id"]}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Perjalanan tidak ditemukan")
    return TripResponse(**trip)

@api_router.put("/trips/{trip_id}", response_model=TripResponse)
async def update_trip(trip_id: str, trip: TripUpdate, current_user: dict = Depends(get_current_user)):
    existing = await db.trips.find_one({"id": trip_id, "user_id": current_user["id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Perjalanan tidak ditemukan")
    
    update_data = {k: v for k, v in trip.model_dump().items() if v is not None}
    if update_data:
        await db.trips.update_one({"id": trip_id}, {"$set": update_data})
    
    updated = await db.trips.find_one({"id": trip_id}, {"_id": 0})
    return TripResponse(**updated)

@api_router.delete("/trips/{trip_id}")
async def delete_trip(trip_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.trips.delete_one({"id": trip_id, "user_id": current_user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Perjalanan tidak ditemukan")
    
    # Delete related itineraries and expenses
    await db.itineraries.delete_many({"trip_id": trip_id})
    await db.expenses.delete_many({"trip_id": trip_id})
    
    return {"message": "Perjalanan berhasil dihapus"}

# ============ ITINERARY ROUTES ============

@api_router.post("/trips/{trip_id}/itineraries", response_model=ItineraryResponse)
async def create_itinerary(trip_id: str, itinerary: ItineraryCreate, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id, "user_id": current_user["id"]})
    if not trip:
        raise HTTPException(status_code=404, detail="Perjalanan tidak ditemukan")
    
    itinerary_id = str(uuid.uuid4())
    itinerary_doc = {
        "id": itinerary_id,
        "trip_id": trip_id,
        "tanggal": itinerary.tanggal,
        "waktu": itinerary.waktu,
        "kegiatan": itinerary.kegiatan,
        "lokasi": itinerary.lokasi,
        "catatan": itinerary.catatan or ""
    }
    await db.itineraries.insert_one(itinerary_doc)
    return ItineraryResponse(**itinerary_doc)

@api_router.get("/trips/{trip_id}/itineraries", response_model=List[ItineraryResponse])
async def get_itineraries(trip_id: str, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id, "user_id": current_user["id"]})
    if not trip:
        raise HTTPException(status_code=404, detail="Perjalanan tidak ditemukan")
    
    itineraries = await db.itineraries.find({"trip_id": trip_id}, {"_id": 0}).to_list(1000)
    # Sort by date and time
    itineraries.sort(key=lambda x: (x["tanggal"], x["waktu"]))
    return [ItineraryResponse(**i) for i in itineraries]

@api_router.put("/trips/{trip_id}/itineraries/{itinerary_id}", response_model=ItineraryResponse)
async def update_itinerary(trip_id: str, itinerary_id: str, itinerary: ItineraryUpdate, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id, "user_id": current_user["id"]})
    if not trip:
        raise HTTPException(status_code=404, detail="Perjalanan tidak ditemukan")
    
    existing = await db.itineraries.find_one({"id": itinerary_id, "trip_id": trip_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Itinerary tidak ditemukan")
    
    update_data = {k: v for k, v in itinerary.model_dump().items() if v is not None}
    if update_data:
        await db.itineraries.update_one({"id": itinerary_id}, {"$set": update_data})
    
    updated = await db.itineraries.find_one({"id": itinerary_id}, {"_id": 0})
    return ItineraryResponse(**updated)

@api_router.delete("/trips/{trip_id}/itineraries/{itinerary_id}")
async def delete_itinerary(trip_id: str, itinerary_id: str, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id, "user_id": current_user["id"]})
    if not trip:
        raise HTTPException(status_code=404, detail="Perjalanan tidak ditemukan")
    
    result = await db.itineraries.delete_one({"id": itinerary_id, "trip_id": trip_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Itinerary tidak ditemukan")
    
    return {"message": "Itinerary berhasil dihapus"}

# ============ EXPENSE ROUTES ============

@api_router.post("/trips/{trip_id}/expenses", response_model=ExpenseResponse)
async def create_expense(trip_id: str, expense: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id, "user_id": current_user["id"]})
    if not trip:
        raise HTTPException(status_code=404, detail="Perjalanan tidak ditemukan")
    
    # Get next number
    count = await db.expenses.count_documents({"trip_id": trip_id})
    
    expense_id = str(uuid.uuid4())
    expense_doc = {
        "id": expense_id,
        "trip_id": trip_id,
        "nomor": count + 1,
        "tanggal": expense.tanggal,
        "uraian": expense.uraian,
        "jumlah": expense.jumlah,
        "catatan": expense.catatan or ""
    }
    await db.expenses.insert_one(expense_doc)
    return ExpenseResponse(**expense_doc)

@api_router.get("/trips/{trip_id}/expenses", response_model=List[ExpenseResponse])
async def get_expenses(trip_id: str, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id, "user_id": current_user["id"]})
    if not trip:
        raise HTTPException(status_code=404, detail="Perjalanan tidak ditemukan")
    
    expenses = await db.expenses.find({"trip_id": trip_id}, {"_id": 0}).sort("nomor", 1).to_list(1000)
    return [ExpenseResponse(**e) for e in expenses]

@api_router.put("/trips/{trip_id}/expenses/{expense_id}", response_model=ExpenseResponse)
async def update_expense(trip_id: str, expense_id: str, expense: ExpenseUpdate, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id, "user_id": current_user["id"]})
    if not trip:
        raise HTTPException(status_code=404, detail="Perjalanan tidak ditemukan")
    
    existing = await db.expenses.find_one({"id": expense_id, "trip_id": trip_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Biaya tidak ditemukan")
    
    update_data = {k: v for k, v in expense.model_dump().items() if v is not None}
    if update_data:
        await db.expenses.update_one({"id": expense_id}, {"$set": update_data})
    
    updated = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    return ExpenseResponse(**updated)

@api_router.delete("/trips/{trip_id}/expenses/{expense_id}")
async def delete_expense(trip_id: str, expense_id: str, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id, "user_id": current_user["id"]})
    if not trip:
        raise HTTPException(status_code=404, detail="Perjalanan tidak ditemukan")
    
    deleted = await db.expenses.find_one({"id": expense_id, "trip_id": trip_id})
    if not deleted:
        raise HTTPException(status_code=404, detail="Biaya tidak ditemukan")
    
    await db.expenses.delete_one({"id": expense_id})
    
    # Renumber remaining expenses
    remaining = await db.expenses.find({"trip_id": trip_id}, {"_id": 0}).sort("nomor", 1).to_list(1000)
    for i, exp in enumerate(remaining):
        await db.expenses.update_one({"id": exp["id"]}, {"$set": {"nomor": i + 1}})
    
    return {"message": "Biaya berhasil dihapus"}

# ============ REPORT ROUTES ============

@api_router.get("/trips/{trip_id}/report/validate")
async def validate_report(trip_id: str, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id, "user_id": current_user["id"]}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Perjalanan tidak ditemukan")
    
    itineraries = await db.itineraries.find({"trip_id": trip_id}, {"_id": 0}).to_list(1000)
    expenses = await db.expenses.find({"trip_id": trip_id}, {"_id": 0}).to_list(1000)
    
    profile_completed = bool(current_user.get("nip") and current_user.get("jabatan") and current_user.get("unit"))
    trip_completed = bool(trip.get("judul") and trip.get("tujuan") and trip.get("tanggal_mulai") and 
                         trip.get("tanggal_selesai") and trip.get("dasar_perjalanan") and trip.get("maksud_tujuan"))
    has_itinerary = len(itineraries) > 0
    has_expense = len(expenses) > 0
    
    return {
        "profile_completed": profile_completed,
        "trip_completed": trip_completed,
        "has_itinerary": has_itinerary,
        "has_expense": has_expense,
        "can_generate": profile_completed and trip_completed and has_itinerary and has_expense,
        "user": {
            "full_name": current_user.get("full_name", ""),
            "nip": current_user.get("nip", ""),
            "jabatan": current_user.get("jabatan", ""),
            "unit": current_user.get("unit", "")
        },
        "trip": trip,
        "itineraries": sorted(itineraries, key=lambda x: (x["tanggal"], x["waktu"])),
        "expenses": sorted(expenses, key=lambda x: x["nomor"]),
        "total_expense": sum(e["jumlah"] for e in expenses)
    }

@api_router.get("/trips/{trip_id}/report")
async def generate_report(trip_id: str, format: str = "pdf", current_user: dict = Depends(get_current_user)):
    validation = await validate_report(trip_id, current_user)
    if not validation["can_generate"]:
        raise HTTPException(status_code=400, detail="Data belum lengkap untuk generate laporan")
    
    if format == "xlsx":
        return await generate_excel_report(validation)
    else:
        return await generate_pdf_report(validation)

async def generate_pdf_report(data: dict):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, alignment=TA_CENTER, spaceAfter=12)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=11, alignment=TA_CENTER, spaceAfter=24)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=11, spaceAfter=8, spaceBefore=16)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=10)
    
    elements = []
    
    # Header
    elements.append(Paragraph("LAPORAN PERJALANAN DINAS", title_style))
    elements.append(Paragraph(f"Nomor: -", subtitle_style))
    
    # Trip Info
    trip = data["trip"]
    user = data["user"]
    
    info_data = [
        ["Nama", ":", user["full_name"]],
        ["NIP", ":", user["nip"]],
        ["Jabatan", ":", user["jabatan"]],
        ["Unit/Bagian", ":", user["unit"]],
        ["", "", ""],
        ["Judul Perjalanan", ":", trip["judul"]],
        ["Tujuan", ":", trip["tujuan"]],
        ["Tanggal", ":", f"{trip['tanggal_mulai']} s.d. {trip['tanggal_selesai']}"],
        ["Dasar Perjalanan", ":", trip["dasar_perjalanan"]],
        ["Maksud dan Tujuan", ":", trip["maksud_tujuan"]],
    ]
    
    info_table = Table(info_data, colWidths=[4*cm, 0.5*cm, 10*cm])
    info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Itinerary Section
    elements.append(Paragraph("I. URAIAN KEGIATAN", heading_style))
    
    itinerary_header = [["No", "Tanggal", "Waktu", "Kegiatan", "Lokasi"]]
    itinerary_data = itinerary_header + [
        [str(i+1), it["tanggal"], it["waktu"], it["kegiatan"], it["lokasi"]]
        for i, it in enumerate(data["itineraries"])
    ]
    
    itinerary_table = Table(itinerary_data, colWidths=[1*cm, 2.5*cm, 1.5*cm, 7*cm, 3*cm])
    itinerary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
    ]))
    elements.append(itinerary_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Expense Section
    elements.append(Paragraph("II. RINCIAN BIAYA", heading_style))
    
    def format_rupiah(amount):
        return f"Rp {amount:,.0f}".replace(",", ".")
    
    expense_header = [["No", "Tanggal", "Uraian", "Jumlah (Rp)"]]
    expense_data = expense_header + [
        [str(exp["nomor"]), exp["tanggal"], exp["uraian"], format_rupiah(exp["jumlah"])]
        for exp in data["expenses"]
    ]
    expense_data.append(["", "", "TOTAL", format_rupiah(data["total_expense"])])
    
    expense_table = Table(expense_data, colWidths=[1*cm, 2.5*cm, 8*cm, 3.5*cm])
    expense_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (2, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.95, 0.95, 0.95)),
    ]))
    elements.append(expense_table)
    elements.append(Spacer(1, 1.5*cm))
    
    # Signature Section
    from datetime import datetime
    today = datetime.now().strftime("%d-%m-%Y")
    
    signature_data = [
        ["", f"__________, {today}"],
        ["Mengetahui,", "Yang Membuat Laporan,"],
        ["Atasan Langsung", ""],
        ["", ""],
        ["", ""],
        ["", ""],
        ["(_________________)", f"({user['full_name']})"],
        ["NIP.", f"NIP. {user['nip']}"],
    ]
    
    sig_table = Table(signature_data, colWidths=[7.5*cm, 7.5*cm])
    sig_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(sig_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"laporan_perjalanan_{trip['judul'].replace(' ', '_')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

async def generate_excel_report(data: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Perjalanan"
    
    trip = data["trip"]
    user = data["user"]
    
    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "LAPORAN PERJALANAN DINAS"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Info Section
    info = [
        ("Nama", user["full_name"]),
        ("NIP", user["nip"]),
        ("Jabatan", user["jabatan"]),
        ("Unit/Bagian", user["unit"]),
        ("", ""),
        ("Judul Perjalanan", trip["judul"]),
        ("Tujuan", trip["tujuan"]),
        ("Tanggal", f"{trip['tanggal_mulai']} s.d. {trip['tanggal_selesai']}"),
        ("Dasar Perjalanan", trip["dasar_perjalanan"]),
        ("Maksud dan Tujuan", trip["maksud_tujuan"]),
    ]
    
    row = 3
    for label, value in info:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = ":"
        ws[f'C{row}'] = value
        row += 1
    
    # Itinerary Section
    row += 1
    ws[f'A{row}'] = "I. URAIAN KEGIATAN"
    ws[f'A{row}'].font = header_font
    row += 1
    
    headers = ["No", "Tanggal", "Waktu", "Kegiatan", "Lokasi"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for i, it in enumerate(data["itineraries"], 1):
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=it["tanggal"]).border = thin_border
        ws.cell(row=row, column=3, value=it["waktu"]).border = thin_border
        ws.cell(row=row, column=4, value=it["kegiatan"]).border = thin_border
        ws.cell(row=row, column=5, value=it["lokasi"]).border = thin_border
        row += 1
    
    # Expense Section
    row += 1
    ws[f'A{row}'] = "II. RINCIAN BIAYA"
    ws[f'A{row}'].font = header_font
    row += 1
    
    headers = ["No", "Tanggal", "Uraian", "Jumlah (Rp)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for exp in data["expenses"]:
        ws.cell(row=row, column=1, value=exp["nomor"]).border = thin_border
        ws.cell(row=row, column=2, value=exp["tanggal"]).border = thin_border
        ws.cell(row=row, column=3, value=exp["uraian"]).border = thin_border
        cell = ws.cell(row=row, column=4, value=exp["jumlah"])
        cell.border = thin_border
        cell.number_format = '#,##0'
        row += 1
    
    # Total
    ws.cell(row=row, column=3, value="TOTAL").font = header_font
    ws.cell(row=row, column=3).border = thin_border
    total_cell = ws.cell(row=row, column=4, value=data["total_expense"])
    total_cell.font = header_font
    total_cell.border = thin_border
    total_cell.number_format = '#,##0'
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"laporan_perjalanan_{trip['judul'].replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============ HEALTH CHECK ============

@api_router.get("/health")
async def health_check():
    return {"status": "healthy"}

# Include router and middleware
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
